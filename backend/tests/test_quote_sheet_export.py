from __future__ import annotations

import csv
import json
from contextlib import redirect_stdout
from io import BytesIO, StringIO
from urllib.parse import unquote

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from drawing_fixtures import PNG_1X1
from helpers import create_admin, create_factory, create_quoter, insert_quote_sheet_template, login
from quote_assistant.domain.part_family import EXPERIMENTAL_MARK_TEXT
from quote_assistant.domain.quality import LOW_QUALITY_MARK_TEXT
from quote_assistant.domain.quote_sheet import (
    DEFAULT_EXPERIMENTAL_MARK_HEADER,
    DEFAULT_LOW_QUALITY_MARK_HEADER,
    DEFAULT_RISK_LABELS_HEADER,
    HIDDEN_UNREVIEWED_EXPORT_MESSAGE,
    default_quote_sheet_template,
)
from quote_assistant.interface.cli.quote_sheet_template import main as quote_sheet_template_cli


def _login_quoter(client: TestClient, db_session: Session) -> None:
    factory_id = create_factory(db_session, "华东精密")
    create_quoter(db_session, factory_id, "quoter_a", "secret-a")
    db_session.commit()
    assert login(client, "quoter_a", "secret-a").status_code == 200


def _upload(client: TestClient, filename: str) -> dict:
    response = client.post(
        "/part-drawings",
        files=[("files", (filename, PNG_1X1, "image/png"))],
    )
    assert response.status_code == 200
    return response.json()["items"][0]


def _create_task(client: TestClient, name: str, customer_name: str) -> dict:
    response = client.post(
        "/quote-tasks",
        json={"name": name, "customer_name": customer_name},
    )
    assert response.status_code == 200
    return response.json()


def _assign(client: TestClient, task_id: str, drawing_id: str) -> None:
    assigned = client.post(
        f"/quote-tasks/{task_id}/part-drawings",
        json={"part_drawing_id": drawing_id},
    )
    assert assigned.status_code == 200


def _switch(client: TestClient, username: str, password: str) -> None:
    client.post("/auth/logout")
    assert login(client, username, password).status_code == 200


def _complete_review(client: TestClient, drawing_id: str) -> None:
    detail = client.get(f"/part-drawings/{drawing_id}")
    assert detail.status_code == 200
    for field in detail.json()["extracted_fields"]:
        if field["requires_confirmation"] and not field["confirmed"] and not field["ignored"]:
            confirmed = client.post(f"/part-drawings/{drawing_id}/fields/{field['key']}/confirm")
            assert confirmed.status_code == 200
    done = client.post(f"/part-drawings/{drawing_id}/complete-review")
    assert done.status_code == 200
    assert done.json()["status"] == "已复核"


def _read_xlsx(content: bytes) -> tuple[list[str], list[list[str]]]:
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active
    assert sheet is not None
    raw_rows = list(sheet.iter_rows(values_only=True))
    assert raw_rows
    headers = ["" if cell is None else str(cell) for cell in raw_rows[0]]
    rows = [["" if cell is None else str(cell) for cell in row] for row in raw_rows[1:]]
    return headers, rows


def _read_csv(content: bytes) -> tuple[list[str], list[list[str]]]:
    reader = csv.reader(StringIO(content.decode("utf-8-sig")))
    raw_rows = list(reader)
    assert raw_rows
    return raw_rows[0], raw_rows[1:]


def test_导出底稿行数等于零件数且含风险标签与实验性列(
    client: TestClient, db_session: Session
) -> None:
    _login_quoter(client, db_session)
    # WIRE-RL-01 drives risk labels; FX-T slot makes it 目标零件族 so 实验性 stays empty.
    risk_drawing = _upload(client, "WIRE-RL-01-FX-TQ.png")
    experimental_drawing = _upload(client, "FX-NQ-01.png")
    task = _create_task(client, "八月询价", "南方汽配")
    _assign(client, task["id"], risk_drawing["id"])
    _assign(client, task["id"], experimental_drawing["id"])
    _complete_review(client, risk_drawing["id"])
    _complete_review(client, experimental_drawing["id"])

    exported = client.get(f"/quote-tasks/{task['id']}/quote-sheet")
    assert exported.status_code == 200
    assert "spreadsheetml" in exported.headers["content-type"]
    disposition = unquote(exported.headers["content-disposition"])
    assert "八月询价-报价底稿.xlsx" in disposition

    headers, rows = _read_xlsx(exported.content)
    assert len(rows) == 2
    assert DEFAULT_RISK_LABELS_HEADER in headers
    assert DEFAULT_EXPERIMENTAL_MARK_HEADER in headers
    assert DEFAULT_LOW_QUALITY_MARK_HEADER in headers
    risk_col = headers.index(DEFAULT_RISK_LABELS_HEADER)
    experimental_col = headers.index(DEFAULT_EXPERIMENTAL_MARK_HEADER)
    drawing_no_col = headers.index("图号")

    by_drawing_no = {row[drawing_no_col]: row for row in rows}
    assert set(by_drawing_no) == {"RL-WIRE-01", "FL-001"}
    assert "高精度" in by_drawing_no["RL-WIRE-01"][risk_col]
    assert "深孔" in by_drawing_no["RL-WIRE-01"][risk_col]
    assert by_drawing_no["RL-WIRE-01"][experimental_col] == ""
    assert by_drawing_no["FL-001"][experimental_col] == EXPERIMENTAL_MARK_TEXT
    assert by_drawing_no["FL-001"][risk_col] == ""
    assert "无风险" not in by_drawing_no["FL-001"][risk_col]
    assert "安全" not in by_drawing_no["FL-001"][risk_col]


def test_csv导出同样一行一件且带风险与实验性列(
    client: TestClient, db_session: Session
) -> None:
    _login_quoter(client, db_session)
    first = _upload(client, "WIRE-RL-01-FX-TQ.png")
    second = _upload(client, "FX-NQ-01.png")
    task = _create_task(client, "八月询价", "南方汽配")
    _assign(client, task["id"], first["id"])
    _assign(client, task["id"], second["id"])
    _complete_review(client, first["id"])
    _complete_review(client, second["id"])

    exported = client.get(f"/quote-tasks/{task['id']}/quote-sheet", params={"format": "csv"})
    assert exported.status_code == 200
    assert "text/csv" in exported.headers["content-type"]
    headers, rows = _read_csv(exported.content)
    assert len(rows) == 2
    assert DEFAULT_RISK_LABELS_HEADER in headers
    assert DEFAULT_EXPERIMENTAL_MARK_HEADER in headers
    experimental_col = headers.index(DEFAULT_EXPERIMENTAL_MARK_HEADER)
    assert EXPERIMENTAL_MARK_TEXT in {row[experimental_col] for row in rows}


def test_未复核零件图拦截导出并点名还差哪几个(
    client: TestClient, db_session: Session
) -> None:
    _login_quoter(client, db_session)
    reviewed = _upload(client, "FX-TQ-01.png")
    pending = _upload(client, "FX-TA-01.png")
    task = _create_task(client, "混装询价", "东海法兰")
    _assign(client, task["id"], reviewed["id"])
    _assign(client, task["id"], pending["id"])
    _complete_review(client, reviewed["id"])

    blocked = client.get(f"/quote-tasks/{task['id']}/quote-sheet")
    assert blocked.status_code == 409
    detail = blocked.json()["detail"]
    assert "未完成复核" in detail
    assert pending["original_filename"] in detail
    assert reviewed["original_filename"] not in detail


def test_列与顺序读取该工厂的模板配置(client: TestClient, db_session: Session) -> None:
    factory_id = create_factory(db_session, "华东精密")
    create_quoter(db_session, factory_id, "quoter_a", "secret-a")
    insert_quote_sheet_template(
        db_session,
        factory_id,
        [
            ("part_name", "品名"),
            ("drawing_no", "本厂图号"),
            ("risk_labels", "加工风险"),
        ],
    )
    db_session.commit()
    assert login(client, "quoter_a", "secret-a").status_code == 200

    drawing = _upload(client, "WIRE-RL-01-FX-TQ.png")
    task = _create_task(client, "定制底稿", "南方汽配")
    _assign(client, task["id"], drawing["id"])
    _complete_review(client, drawing["id"])

    exported = client.get(f"/quote-tasks/{task['id']}/quote-sheet")
    assert exported.status_code == 200
    headers, rows = _read_xlsx(exported.content)
    assert headers == [
        "品名",
        "本厂图号",
        "加工风险",
        DEFAULT_EXPERIMENTAL_MARK_HEADER,
        DEFAULT_LOW_QUALITY_MARK_HEADER,
    ]
    assert len(rows) == 1
    assert rows[0][0] == "风险接线件"
    assert rows[0][1] == "RL-WIRE-01"
    assert "高精度" in rows[0][2]


def test_导出包含低质量图标记(client: TestClient, db_session: Session) -> None:
    _login_quoter(client, db_session)
    drawing = _upload(client, "FX-TP-01.png")
    continued = client.post(f"/part-drawings/{drawing['id']}/continue-despite-quality")
    assert continued.status_code == 200
    assert continued.json()["low_quality_mark"] == LOW_QUALITY_MARK_TEXT
    task = _create_task(client, "差图询价", "东海法兰")
    _assign(client, task["id"], drawing["id"])
    _complete_review(client, drawing["id"])

    exported = client.get(f"/quote-tasks/{task['id']}/quote-sheet")
    assert exported.status_code == 200
    headers, rows = _read_xlsx(exported.content)
    mark_col = headers.index(DEFAULT_LOW_QUALITY_MARK_HEADER)
    assert len(rows) == 1
    assert rows[0][mark_col] == LOW_QUALITY_MARK_TEXT


def test_没有字段映射的HTTP入口(client: TestClient, db_session: Session) -> None:
    _login_quoter(client, db_session)
    assert client.get("/quote-sheet-templates").status_code == 404
    assert client.post("/quote-sheet-templates", json={"columns": []}).status_code == 404
    assert client.get("/admin/quote-sheet-template").status_code == 404


def test_跨报价员未复核零件拦住导出且不泄漏文件名(
    client: TestClient, db_session: Session
) -> None:
    factory_id = create_factory(db_session, "华东精密")
    create_quoter(db_session, factory_id, "quoter_a", "secret-a")
    create_quoter(db_session, factory_id, "quoter_c", "secret-c")
    create_admin(db_session, factory_id, "admin_a", "secret-admin")
    db_session.commit()

    assert login(client, "quoter_a", "secret-a").status_code == 200
    drawing_a = _upload(client, "FX-TQ-01.png")
    task = _create_task(client, "混装询价", "东海法兰")
    _assign(client, task["id"], drawing_a["id"])
    _complete_review(client, drawing_a["id"])

    _switch(client, "quoter_c", "secret-c")
    drawing_c = _upload(client, "C报价员-未复核-轴.png")

    _switch(client, "admin_a", "secret-admin")
    assigned = client.post(
        f"/quote-tasks/{task['id']}/part-drawings",
        json={"part_drawing_id": drawing_c["id"]},
    )
    assert assigned.status_code == 200

    _switch(client, "quoter_a", "secret-a")
    detail = client.get(f"/quote-tasks/{task['id']}")
    assert detail.status_code == 200
    body = detail.json()
    assert body["unreviewed_member_count"] == 1
    assert [row["id"] for row in body["drawings"]] == [drawing_a["id"]]
    assert drawing_c["original_filename"] not in json.dumps(body, ensure_ascii=False)

    blocked = client.get(f"/quote-tasks/{task['id']}/quote-sheet")
    assert blocked.status_code == 409
    message = blocked.json()["detail"]
    assert message == HIDDEN_UNREVIEWED_EXPORT_MESSAGE.format(count=1)
    assert drawing_c["original_filename"] not in message

    _switch(client, "quoter_c", "secret-c")
    _complete_review(client, drawing_c["id"])

    _switch(client, "quoter_a", "secret-a")
    ready = client.get(f"/quote-tasks/{task['id']}")
    assert ready.json()["unreviewed_member_count"] == 0
    exported = client.get(f"/quote-tasks/{task['id']}/quote-sheet")
    assert exported.status_code == 200
    _headers, rows = _read_xlsx(exported.content)
    assert len(rows) == 1


def test_cli写入模板后该厂导出列一致未写入的厂用默认列(
    client: TestClient, db_session: Session, database_url: str, tmp_path
) -> None:
    factory_custom = create_factory(db_session, "华东精密")
    factory_default = create_factory(db_session, "南方模具")
    create_quoter(db_session, factory_custom, "quoter_a", "secret-a")
    create_quoter(db_session, factory_default, "quoter_b", "secret-b")
    db_session.commit()

    columns_file = tmp_path / "huadong.json"
    columns_file.write_text(
        json.dumps(
            [
                {"source_key": "part_name", "header": "品名"},
                {"source_key": "drawing_no", "header": "本厂图号"},
                {"source_key": "risk_labels", "header": "加工风险"},
            ],
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    assert (
        quote_sheet_template_cli(
            ["save", "--factory-name", "华东精密", "--columns-file", str(columns_file)],
            database_url=database_url,
        )
        == 0
    )
    assert (
        quote_sheet_template_cli(
            ["save", "--factory-name", "华东精密", "--column", "not_a_field:假列"],
            database_url=database_url,
        )
        == 1
    )

    shown = StringIO()
    with redirect_stdout(shown):
        assert (
            quote_sheet_template_cli(
                ["show", "--factory-name", "南方模具", "--json"],
                database_url=database_url,
            )
            == 0
        )
    default_show = json.loads(shown.getvalue())
    assert default_show["stored"] is False
    assert [column["header"] for column in default_show["resolved_columns"]] == [
        column.header for column in default_quote_sheet_template().columns
    ]

    assert login(client, "quoter_a", "secret-a").status_code == 200
    drawing_custom = _upload(client, "WIRE-RL-01-FX-TQ.png")
    task_custom = _create_task(client, "定制底稿", "南方汽配")
    _assign(client, task_custom["id"], drawing_custom["id"])
    _complete_review(client, drawing_custom["id"])
    exported_custom = client.get(f"/quote-tasks/{task_custom['id']}/quote-sheet")
    assert exported_custom.status_code == 200
    headers_custom, rows_custom = _read_xlsx(exported_custom.content)
    assert headers_custom == [
        "品名",
        "本厂图号",
        "加工风险",
        DEFAULT_EXPERIMENTAL_MARK_HEADER,
        DEFAULT_LOW_QUALITY_MARK_HEADER,
    ]
    assert len(rows_custom) == 1

    _switch(client, "quoter_b", "secret-b")
    drawing_default = _upload(client, "WIRE-RL-01-FX-TQ.png")
    task_default = _create_task(client, "默认底稿", "南方汽配")
    _assign(client, task_default["id"], drawing_default["id"])
    _complete_review(client, drawing_default["id"])
    exported_default = client.get(f"/quote-tasks/{task_default['id']}/quote-sheet")
    assert exported_default.status_code == 200
    headers_default, rows_default = _read_xlsx(exported_default.content)
    assert headers_default == [column.header for column in default_quote_sheet_template().columns]
    assert len(rows_default) == 1
